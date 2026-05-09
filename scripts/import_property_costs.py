#!/usr/bin/env python3
"""
property_costs 테이블 일괄 임포트
원본: 월세납부리스트.xlsx + 관리비납부.xlsx + 부동산 계약 정보.xlsx
대상: heiro-dev.property_costs

실행: python3 scripts/import_property_costs.py [--dry-run]
"""

import sys
import os
import re
import json
import openpyxl
import mysql.connector
from datetime import datetime

# ── DB 설정 ──────────────────────────────────────────────
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "dev-db.c98404c20038.ap-northeast-2.rds.amazonaws.com"),
    "port": int(os.getenv("DB_PORT", 3306)),
    "user": os.getenv("DB_USER", "admin"),
    "password": os.getenv("DB_PASSWORD", "MRM7pmqtQoWnXMt4lW5j"),
    "database": os.getenv("DB_NAME", "heiro-dev"),
}

DRY_RUN = "--dry-run" in sys.argv

# ── 파일 경로 ────────────────────────────────────────────
RENT_FILE = os.path.expanduser("~/Downloads/월세납부리스트.xlsx")
MGMT_FILE = os.path.expanduser("~/Downloads/관리비납부.xlsx")
CONTRACT_FILE = os.path.expanduser("~/Downloads/부동산 계약 정보.xlsx")

# ── 숙소명 정규화 (엑셀 → DB 매칭) ───────────────────────
ALIASES = {
    "큐브1차": "청광1차",
    "큐브3차": "청광3차",
    "렘브란트": "렘브",
    "오금스타": "송파오금",
    "미사 마이움": "미사마이움",
    "미사롯데": "미사롯데",
    "미사힐스": "미사힐스",
    "다성": "길동다성",
    "강동큐브": "강동Qv",
    "강동큐브 2차": "대치",  # SK D&D 623 = 대치 101 (V11)
}


def normalize_name(raw_name):
    """엑셀 숙소명 → DB name에서 검색할 키워드 추출"""
    if not raw_name:
        return None, None

    raw_name = raw_name.strip()

    # "예건 902 최익진" → "예건", "902"
    # "센텀2차 301" → "센텀2차", "301"
    # 소유자 이름이나 부가정보 제거
    parts = raw_name.split()
    if len(parts) < 2:
        return None, None

    building = parts[0]
    room = parts[1]

    # 별칭 변환
    for alias, replacement in ALIASES.items():
        if building == alias:
            building = replacement
            break

    # 호수에서 숫자만 추출 (예: "102-803" 그대로, "301" 그대로)
    return building, room


def load_properties_from_db(conn):
    """DB에서 properties 조회, 매칭용 딕셔너리 생성"""
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, code, name, building_name, monthly_rent FROM properties")
    rows = cursor.fetchall()
    cursor.close()
    return rows


def match_property(properties, building, room):
    """building + room으로 property 매칭"""
    candidates = []
    for p in properties:
        name = (p["name"] or "").lower()
        code = (p["code"] or "").lower()

        # name에 building + room이 포함되는지 확인
        b_lower = building.lower()
        r_lower = room.lower()

        if b_lower in name and r_lower in name:
            candidates.append(p)
        elif b_lower in code and r_lower in code:
            candidates.append(p)

    if len(candidates) == 1:
        return candidates[0]
    elif len(candidates) > 1:
        # 가장 정확한 매칭 (이름이 짧은 것 우선)
        candidates.sort(key=lambda x: len(x["name"] or ""))
        return candidates[0]
    return None


def parse_rent_excel():
    """월세납부리스트.xlsx 파싱"""
    wb = openpyxl.load_workbook(RENT_FILE, data_only=True)
    ws = wb["Sheet1"]
    results = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        name_val = row[1].value
        rent_val = row[2].value
        holder = row[3].value
        bank = row[4].value
        account = row[5].value

        if not name_val or not rent_val:
            continue
        if "총합" in str(name_val) or "수수료" in str(name_val):
            continue

        building, room = normalize_name(str(name_val))
        if not building or not room:
            continue

        # 만 단위 → 원 단위
        rent_won = int(float(rent_val) * 10000)

        results.append({
            "raw_name": str(name_val).strip(),
            "building": building,
            "room": room,
            "rent": rent_won,
            "rent_recipient": str(holder).strip() if holder else "",
            "bank": str(bank).strip() if bank else "",
            "account": str(account).strip() if account else "",
        })

    return results


def parse_mgmt_excel():
    """관리비납부.xlsx 파싱 → 숙소별 월평균 관리비"""
    wb = openpyxl.load_workbook(MGMT_FILE, data_only=True)
    ws = wb["Sheet1"]

    # 4개월분 데이터 (1월~4월), 8열씩
    # 각 블록: 숙소 | 관리비 | 가스 | 전기 | 인터넷 | 수도 | 합계 | (빈칸)
    results = {}

    for month_idx in range(4):  # 0=1월, 1=2월, 2=3월, 3=4월
        col_start = month_idx * 8  # 0, 8, 16, 24

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
            name_cell = row[col_start].value if col_start < len(row) else None
            if not name_cell:
                continue

            name_str = str(name_cell).strip()
            if not name_str or name_str == "숙소":
                continue

            # 각 항목 추출
            mgmt = row[col_start + 1].value if col_start + 1 < len(row) else 0
            gas = row[col_start + 2].value if col_start + 2 < len(row) else 0
            electric = row[col_start + 3].value if col_start + 3 < len(row) else 0
            internet = row[col_start + 4].value if col_start + 4 < len(row) else 0
            water = row[col_start + 5].value if col_start + 5 < len(row) else 0

            # 숙소명에서 "호" 제거: "가산지웰 328호" → "가산지웰 328"
            name_clean = name_str.replace("호", "").strip()

            if name_clean not in results:
                results[name_clean] = {
                    "months": 0,
                    "management_fee": 0,
                    "gas": 0,
                    "electric": 0,
                    "internet": 0,
                    "water": 0,
                }

            total_this_month = sum(int(v or 0) for v in [mgmt, gas, electric, internet, water])
            if total_this_month > 0:
                results[name_clean]["months"] += 1
                results[name_clean]["management_fee"] += int(mgmt or 0)
                results[name_clean]["gas"] += int(gas or 0)
                results[name_clean]["electric"] += int(electric or 0)
                results[name_clean]["internet"] += int(internet or 0)
                results[name_clean]["water"] += int(water or 0)

    # 월평균 계산
    averaged = {}
    for name, data in results.items():
        months = data["months"]
        if months == 0:
            continue

        building_room = name.split()
        if len(building_room) < 2:
            continue

        building = building_room[0]
        room = building_room[1]

        # 별칭 변환
        for alias, replacement in ALIASES.items():
            if building == alias:
                building = replacement
                break

        key = f"{building}_{room}"
        averaged[key] = {
            "raw_name": name,
            "building": building,
            "room": room,
            "management_fee": data["management_fee"] // months,
            "gas": data["gas"] // months,
            "electric": data["electric"] // months,
            "internet": data["internet"] // months,
            "water": data["water"] // months,
        }

    return averaged


def parse_contract_excel():
    """부동산 계약 정보.xlsx 파싱 → 보증금, 계약기간"""
    wb = openpyxl.load_workbook(CONTRACT_FILE, data_only=True)
    ws = wb["Sheet1"]
    results = {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        name_val = row[1].value  # 숙소명
        owner_type = row[2].value  # 유형
        deposit = row[4].value  # 보증금
        rent = row[5].value  # 임대료
        contract_period = row[10].value  # 계약기간

        if not name_val:
            continue

        name_str = str(name_val).strip()
        parts = name_str.split()
        if len(parts) < 2:
            continue

        building = parts[0]
        room = parts[1].replace("호", "")

        for alias, replacement in ALIASES.items():
            if building == alias:
                building = replacement
                break

        key = f"{building}_{room}"
        results[key] = {
            "raw_name": name_str,
            "deposit": int(float(deposit)) if deposit and str(deposit).replace('.','').replace('-','').isdigit() else 0,
            "contract_rent": int(float(rent)) if rent and str(rent).replace('.','').replace('-','').isdigit() else 0,
            "owner_type": str(owner_type).strip() if owner_type else "단기임대",
            "contract_period": str(contract_period).strip() if contract_period else "",
        }

    return results


def build_utilities_json(mgmt_data):
    """관리비 데이터 → JSON utilities 구조"""
    if not mgmt_data:
        return json.dumps({
            "management_fee": {"mode": "VARIABLE", "amount": 0},
            "internet": {"mode": "FIXED", "amount": 0},
            "electric": {"mode": "VARIABLE", "amount": 0},
            "gas": {"mode": "VARIABLE", "amount": 0},
            "water": {"mode": "VARIABLE", "amount": 0},
            "insurance": {"mode": "FIXED", "amount": 0},
            "other_utility": {"mode": "FIXED", "amount": 0},
        })

    return json.dumps({
        "management_fee": {"mode": "VARIABLE", "amount": mgmt_data.get("management_fee", 0)},
        "internet": {"mode": "FIXED", "amount": mgmt_data.get("internet", 0)},
        "electric": {"mode": "VARIABLE", "amount": mgmt_data.get("electric", 0)},
        "gas": {"mode": "VARIABLE", "amount": mgmt_data.get("gas", 0)},
        "water": {"mode": "VARIABLE", "amount": mgmt_data.get("water", 0)},
        "insurance": {"mode": "FIXED", "amount": 0},
        "other_utility": {"mode": "FIXED", "amount": 0},
    })


def main():
    print("=" * 70)
    print("HIERO property_costs 일괄 임포트")
    print(f"모드: {'DRY-RUN (DB 변경 없음)' if DRY_RUN else 'LIVE (DB에 INSERT)'}")
    print("=" * 70)

    # 1. 엑셀 파싱
    print("\n[1/5] 월세납부리스트 파싱...")
    rent_data = parse_rent_excel()
    print(f"  → {len(rent_data)}개 숙소 월세 추출")

    print("[2/5] 관리비납부 파싱...")
    mgmt_data = parse_mgmt_excel()
    print(f"  → {len(mgmt_data)}개 숙소 관리비 추출")

    print("[3/5] 부동산 계약 정보 파싱...")
    contract_data = parse_contract_excel()
    print(f"  → {len(contract_data)}개 숙소 계약 추출")

    # 2. DB 연결 + properties 조회
    print("\n[4/5] DB 연결 및 properties 조회...")
    conn = mysql.connector.connect(**DB_CONFIG)
    properties = load_properties_from_db(conn)
    print(f"  → {len(properties)}개 properties 로드")

    # 3. 매칭 + INSERT
    print("\n[5/5] 매칭 및 INSERT...")
    matched = 0
    unmatched = []
    duplicates = set()
    insert_rows = []

    for rent in rent_data:
        building = rent["building"]
        room = rent["room"]
        key = f"{building}_{room}"

        # 중복 체크 (엑셀에 같은 숙소 여러 행: 더하임 204 등)
        if key in duplicates:
            continue
        duplicates.add(key)

        prop = match_property(properties, building, room)
        if not prop:
            unmatched.append(f"  ❌ {rent['raw_name']} ({building} {room})")
            continue

        # 관리비 매칭
        mgmt = mgmt_data.get(key, None)
        # 계약 매칭
        contract = contract_data.get(key, None)

        # 소유구조 결정
        owner_type = "LEASED"  # 기본: 임차
        if contract:
            ot = contract.get("owner_type", "")
            if "대위변제" in ot:
                owner_type = "OWNED"
            elif "위탁" in ot:
                owner_type = "CONSIGNED"

        deposit = contract.get("deposit", 0) if contract else 0
        utilities_json = build_utilities_json(mgmt)

        insert_rows.append({
            "property_id": prop["id"],
            "property_name": prop["name"],
            "raw_name": rent["raw_name"],
            "owner_type": owner_type,
            "rent": rent["rent"],
            "rent_recipient": rent["rent_recipient"],
            "deposit": deposit,
            "utilities": utilities_json,
        })
        matched += 1

    # 결과 출력
    print(f"\n{'=' * 70}")
    print(f"매칭 결과: {matched}개 성공 / {len(unmatched)}개 실패")
    print(f"{'=' * 70}")

    if unmatched:
        print("\n미매칭 목록:")
        for u in unmatched:
            print(u)

    print(f"\n매칭 성공 목록:")
    for r in insert_rows:
        mgmt_total = 0
        try:
            utils = json.loads(r["utilities"])
            mgmt_total = sum(item["amount"] for item in utils.values())
        except:
            pass
        print(f"  ✅ [{r['property_id']:>3}] {r['property_name'][:35]:<35} "
              f"월세: {r['rent']:>10,}원  관리비(월평균): {mgmt_total:>8,}원  "
              f"보증금: {r['deposit']:>10,}원  {r['owner_type']}")

    if DRY_RUN:
        print(f"\n🔸 DRY-RUN 모드: DB 변경 없음. 실행하려면 --dry-run 제거")
        conn.close()
        return

    # INSERT
    print(f"\n💾 DB INSERT 시작...")
    cursor = conn.cursor()

    inserted = 0
    updated = 0
    for r in insert_rows:
        try:
            cursor.execute("""
                INSERT INTO property_costs
                (property_id, owner_type, rent, rent_recipient, deposit, utilities, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    owner_type = VALUES(owner_type),
                    rent = VALUES(rent),
                    rent_recipient = VALUES(rent_recipient),
                    deposit = VALUES(deposit),
                    utilities = VALUES(utilities),
                    updated_at = NOW()
            """, (
                r["property_id"],
                r["owner_type"],
                r["rent"],
                r["rent_recipient"],
                r["deposit"],
                r["utilities"],
            ))
            if cursor.lastrowid:
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            print(f"  ⚠️ [{r['property_id']}] {r['property_name']}: {e}")

    conn.commit()
    cursor.close()
    conn.close()

    print(f"\n✅ 완료: {inserted}개 INSERT, {updated}개 UPDATE")


if __name__ == "__main__":
    main()
