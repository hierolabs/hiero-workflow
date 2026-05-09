#!/usr/bin/env python3
"""
investors + property_investors + property_parkings 일괄 임포트
원본: 월세납부리스트.xlsx + 부동산 계약 정보.xlsx
"""

import os
import sys
import re
import openpyxl
import mysql.connector

DB_CONFIG = {
    "host": "dev-db.c98404c20038.ap-northeast-2.rds.amazonaws.com",
    "port": 3306, "user": "admin",
    "password": "MRM7pmqtQoWnXMt4lW5j", "database": "heiro-dev",
}

RENT_FILE = os.path.expanduser("~/Downloads/월세납부리스트.xlsx")
CONTRACT_FILE = os.path.expanduser("~/Downloads/부동산 계약 정보.xlsx")

ALIASES = {
    "큐브1차": "청광1차", "큐브3차": "청광3차", "렘브란트": "렘브",
    "오금스타": "송파오금", "미사 마이움": "미사마이움", "다성": "길동다성",
    "강동큐브": "강동Qv",
}


def normalize(raw):
    if not raw: return None, None
    raw = str(raw).strip()
    parts = raw.split()
    if len(parts) < 2: return None, None
    building = parts[0]
    room = parts[1].replace("호", "")
    for alias, rep in ALIASES.items():
        if building == alias:
            building = rep
            break
    return building, room


def match_property(properties, building, room):
    for p in properties:
        name = (p["name"] or "").lower()
        if building.lower() in name and room.lower() in name:
            return p
    return None


def main():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)

    # properties 로드
    cursor.execute("SELECT id, code, name FROM properties")
    properties = cursor.fetchall()

    # === 1. investors 채우기 ===
    print("=" * 60)
    print("[1/3] investors 채우기")
    print("=" * 60)

    # 두 엑셀에서 고유 투자자 추출
    all_investors = {}

    # 월세납부리스트
    wb1 = openpyxl.load_workbook(RENT_FILE, data_only=True)
    ws1 = wb1["Sheet1"]
    for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, values_only=False):
        holder = row[3].value
        bank = row[4].value
        account = row[5].value
        if not holder or str(holder).strip() in ("None", "예금주", "월세 총합"):
            continue
        name = str(holder).strip()
        if name not in all_investors:
            all_investors[name] = {
                "bank": str(bank).strip() if bank else "",
                "account": str(account).strip() if account else "",
            }

    # 부동산 계약 정보
    wb2 = openpyxl.load_workbook(CONTRACT_FILE, data_only=True)
    ws2 = wb2["Sheet1"]
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=False):
        holder = row[7].value
        bank = row[8].value
        account = row[9].value
        if not holder: continue
        name = str(holder).strip()
        if name not in all_investors:
            all_investors[name] = {
                "bank": str(bank).strip() if bank else "",
                "account": str(account).strip() if account else "",
            }

    # 기존 investors 확인
    cursor.execute("SELECT id, name FROM investors")
    existing = {r["name"]: r["id"] for r in cursor.fetchall()}

    investor_map = dict(existing)
    inserted = 0
    for name, info in all_investors.items():
        if name in investor_map:
            continue
        cursor.execute("""
            INSERT INTO investors (name, account_holder, bank_name, account_number, created_at, updated_at)
            VALUES (%s, %s, %s, %s, NOW(), NOW())
        """, (name, name, info["bank"], info["account"]))
        investor_map[name] = cursor.lastrowid
        inserted += 1

    conn.commit()
    print(f"  ✅ {inserted}명 INSERT (기존 {len(existing)}, 총 {len(investor_map)}명)")

    # === 2. property_investors 채우기 ===
    print(f"\n{'=' * 60}")
    print("[2/3] property_investors 채우기")
    print("=" * 60)

    # 월세납부리스트: 숙소 → 예금주(투자자) 매핑
    prop_investor_pairs = []
    for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, values_only=False):
        raw_name = row[1].value
        holder = row[3].value
        rent = row[2].value

        if not raw_name or not holder: continue
        if str(raw_name).strip() in ("숙소명", "월세 총합", "수수료"): continue

        building, room = normalize(str(raw_name))
        if not building: continue

        prop = match_property(properties, building, room)
        if not prop: continue

        inv_name = str(holder).strip()
        if inv_name not in investor_map: continue

        rent_won = int(float(rent) * 10000) if rent else 0

        prop_investor_pairs.append({
            "property_id": prop["id"],
            "investor_id": investor_map[inv_name],
            "ownership_type": "sublease",
            "rent_amount": rent_won,
        })

    # 부동산 계약 정보에서 소유구조 보강
    ownership_map = {}
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=False):
        raw_name = row[1].value
        owner_type = row[2].value
        contract_period = row[10].value
        if not raw_name: continue
        building, room = normalize(str(raw_name))
        if not building: continue
        prop = match_property(properties, building, room)
        if not prop: continue

        ot = "sublease"
        if owner_type and "대위변제" in str(owner_type):
            ot = "direct"
        elif owner_type and "위탁" in str(owner_type):
            ot = "consignment"

        contract_str = str(contract_period).strip() if contract_period else ""
        ownership_map[prop["id"]] = {"type": ot, "contract": contract_str}

    # 중복 제거
    seen = set()
    unique_pairs = []
    for p in prop_investor_pairs:
        key = (p["property_id"], p["investor_id"])
        if key in seen: continue
        seen.add(key)
        om = ownership_map.get(p["property_id"], {})
        p["ownership_type"] = om.get("type", "sublease")

        # 계약기간 파싱
        contract = om.get("contract", "")
        start, end = "", ""
        dates = re.findall(r'\d{4}-\d{2}-\d{2}', contract)
        if len(dates) >= 2:
            start, end = dates[0], dates[1]
        elif len(dates) == 1:
            start = dates[0]

        p["contract_start"] = start
        p["contract_end"] = end
        unique_pairs.append(p)

    cursor.execute("DELETE FROM property_investors")
    pi_count = 0
    for p in unique_pairs:
        cursor.execute("""
            INSERT INTO property_investors
            (property_id, investor_id, ownership_type, contract_start, contract_end, rent_amount, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """, (p["property_id"], p["investor_id"], p["ownership_type"],
              p["contract_start"] or None, p["contract_end"] or None, p["rent_amount"]))
        pi_count += 1

    conn.commit()
    print(f"  ✅ {pi_count}개 숙소-투자자 매핑 INSERT")

    # === 3. property_parkings 채우기 ===
    print(f"\n{'=' * 60}")
    print("[3/3] property_parkings 채우기 (부동산 계약 정보 기반)")
    print("=" * 60)

    # 주소에서 건물명 추출 → 건물별 기본 주차 정보
    cursor.execute("SELECT id, name, building_name, address FROM properties WHERE building_name != ''")
    props_with_building = cursor.fetchall()

    cursor.execute("DELETE FROM property_parkings")
    parking_count = 0
    seen_props = set()

    for prop in props_with_building:
        if prop["id"] in seen_props: continue
        seen_props.add(prop["id"])

        building = prop["building_name"] or ""

        # 건물별 기본 주차 정보 (수동 데이터가 없으므로 기본값)
        self_parking = "선착"  # 대부분 오피스텔은 선착순
        street_parking = "가능"

        cursor.execute("""
            INSERT INTO property_parkings
            (property_id, building_name, self_parking, street_parking, created_at, updated_at)
            VALUES (%s, %s, %s, %s, NOW(), NOW())
        """, (prop["id"], building, self_parking, street_parking))
        parking_count += 1

    conn.commit()
    print(f"  ✅ {parking_count}개 숙소 주차 정보 INSERT")

    # === 검증 ===
    print(f"\n{'=' * 60}")
    print("[검증]")
    print("=" * 60)
    cursor.execute("SELECT COUNT(*) as cnt FROM investors")
    print(f"  investors: {cursor.fetchone()['cnt']}명")
    cursor.execute("SELECT COUNT(*) as cnt FROM property_investors")
    print(f"  property_investors: {cursor.fetchone()['cnt']}개")
    cursor.execute("SELECT COUNT(*) as cnt FROM property_parkings")
    print(f"  property_parkings: {cursor.fetchone()['cnt']}개")

    cursor.execute("""
        SELECT i.name, COUNT(pi.id) as props, SUM(pi.rent_amount) as total_rent
        FROM investors i
        LEFT JOIN property_investors pi ON pi.investor_id = i.id
        GROUP BY i.id, i.name
        HAVING props > 0
        ORDER BY total_rent DESC LIMIT 10
    """)
    print(f"\n  투자자별 숙소 수 (top 10):")
    for r in cursor.fetchall():
        print(f"    {r['name']:<12} {r['props']:>2}개  월세합계: {r['total_rent']:>10,}원")

    cursor.close()
    conn.close()
    print(f"\n✅ 완료")


if __name__ == "__main__":
    main()
